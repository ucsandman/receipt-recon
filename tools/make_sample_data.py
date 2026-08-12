"""Generate a realistic sample expense report + matching receipt PDFs.

The sample set has DELIBERATE, DOCUMENTED defects planted in it, so anyone can
verify an audit tool actually catches real problems instead of just looking busy.

Usage:
    python tools/make_sample_data.py                 # 40 rows, the demo set
    python tools/make_sample_data.py --count 350     # full-scale stress set

Output:
    sample-data/expense-report.xlsx
    sample-data/receipts/*.pdf
    sample-data/EXPECTED-FINDINGS.md
"""

from __future__ import annotations

import argparse
import random
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path

# pypdfium2 (Apache-2.0 / BSD-3) rasterizes a few receipts into scan-only PDFs.
# Deliberately NOT PyMuPDF: PyMuPDF is AGPL-3.0, which would conflict with
# publishing this repo under MIT.
import pypdfium2 as pdfium
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "sample-data"
RECEIPTS = OUT / "receipts"

SEED = 20260812  # fixed so the sample set is byte-stable across regenerations


# --------------------------------------------------------------------------
# Data model
# --------------------------------------------------------------------------


@dataclass
class LineItem:
    label: str
    amount: float


@dataclass
class Txn:
    """One expense line. `sheet_*` is what the employee typed into the report.
    `pdf_*` is what the receipt actually says. When they disagree, that is a
    planted defect."""

    txn_id: str
    employee: str
    sheet_date: date
    sheet_vendor: str
    category: str
    sheet_amount: float
    sheet_currency: str
    purpose: str
    approver: str

    # receipt side
    receipt_file: str | None = None
    pdf_date: date | None = None
    pdf_vendor: str | None = None
    pdf_total: float | None = None
    pdf_currency: str | None = None
    items: list[LineItem] = field(default_factory=list)
    tax: float = 0.0
    tip: float = 0.0
    style: str = "retail"
    scanned: bool = False  # render as image-only PDF -> no text layer, forces OCR

    # bookkeeping for the expected-findings doc
    defects: list[str] = field(default_factory=list)

    def resolved(self, attr: str):
        """Receipt value, defaulting to the sheet value when not overridden."""
        pdf = getattr(self, f"pdf_{attr}")
        return pdf if pdf is not None else getattr(self, f"sheet_{attr}")


EMPLOYEES = [
    ("Dana Whitfield", "M. Okonkwo"),
    ("Priya Raghunathan", "M. Okonkwo"),
    ("Tomas Ferreira", "L. Brennan"),
    ("Aisha Nkemdirim", "L. Brennan"),
    ("Grant Halloway", "M. Okonkwo"),
]

VENDORS = {
    "Meals": [
        ("Bluebird Diner", "restaurant"),
        ("Kanto Ramen House", "restaurant"),
        ("The Copper Kettle", "restaurant"),
        ("Marisol Cantina", "restaurant"),
        ("Sunrise Coffee Roasters", "retail"),
    ],
    "Travel": [
        ("Cascadia Airlines", "airline"),
        ("Harborview Suites", "hotel"),
        ("MetroLift", "rideshare"),
        ("Yellow Star Taxi", "taxi"),
    ],
    "Office Supplies": [
        ("Penfield Office Depot", "retail"),
        ("Cartwright Stationers", "retail"),
    ],
    "Software": [
        ("Nordvane Analytics BV", "invoice"),
        ("Gitspark Cloud Inc.", "invoice"),
    ],
    "Client Entertainment": [
        ("Ashcroft Chophouse", "restaurant"),
        ("The Regency Room", "restaurant"),
    ],
}

PURPOSES = [
    "Client onboarding meeting",
    "Q3 regional review",
    "Vendor negotiation",
    "Team offsite",
    "Prospect lunch",
    "Conference travel",
    "Monthly team supplies",
    "Recurring team license",
]


# --------------------------------------------------------------------------
# Receipt rendering
# --------------------------------------------------------------------------


def _money(v: float, cur: str = "USD") -> str:
    sym = {"USD": "$", "EUR": "€", "GBP": "£", "CAD": "C$"}.get(cur, "")
    return f"{sym}{v:,.2f}"


def render_receipt(t: Txn, path: Path) -> None:
    """Draw a receipt PDF. Layout varies by style so extraction is not trivially
    template-matched, which is exactly the situation a real auditor faces."""
    c = canvas.Canvas(str(path), pagesize=letter)
    w, h = letter
    cur = t.resolved("currency")
    vendor = t.resolved("vendor")
    d = t.resolved("date")
    total = t.pdf_total if t.pdf_total is not None else t.sheet_amount

    y = h - 1.1 * inch

    if t.style in ("restaurant", "retail", "taxi"):
        c.setFont("Helvetica-Bold", 15)
        c.drawCentredString(w / 2, y, vendor.upper())
        y -= 16
        c.setFont("Helvetica", 8.5)
        c.drawCentredString(w / 2, y, "1847 Kestrel Avenue, Portland OR 97209")
        y -= 11
        c.drawCentredString(w / 2, y, "(503) 555-0142")
        y -= 22
        c.setFont("Helvetica", 9.5)
        c.drawString(1.1 * inch, y, f"Date: {d.strftime('%m/%d/%Y')}")
        c.drawRightString(
            w - 1.1 * inch, y, f"Server: {random.choice('ABCDJKMR')}. Lee"
        )
        y -= 12
        c.drawString(1.1 * inch, y, f"Check #{random.randint(10000, 99999)}")
        y -= 18
        c.line(1.1 * inch, y, w - 1.1 * inch, y)
        y -= 16

        for it in t.items:
            c.setFont("Helvetica", 10)
            c.drawString(1.2 * inch, y, it.label)
            c.drawRightString(w - 1.2 * inch, y, _money(it.amount, cur))
            y -= 14

        y -= 4
        c.line(4.6 * inch, y, w - 1.1 * inch, y)
        y -= 15
        sub = sum(i.amount for i in t.items)
        for label, val in (("Subtotal", sub), ("Sales Tax", t.tax), ("Tip", t.tip)):
            if val:
                c.setFont("Helvetica", 10)
                c.drawRightString(w - 2.3 * inch, y, f"{label}:")
                c.drawRightString(w - 1.2 * inch, y, _money(val, cur))
                y -= 14
        c.setFont("Helvetica-Bold", 11.5)
        c.drawRightString(w - 2.3 * inch, y, "TOTAL:")
        c.drawRightString(w - 1.2 * inch, y, _money(total, cur))
        y -= 26
        c.setFont("Helvetica", 8.5)
        c.drawCentredString(
            w / 2, y, f"VISA ****{random.randint(1000, 9999)}   APPROVED"
        )
        y -= 11
        c.drawCentredString(w / 2, y, "THANK YOU FOR YOUR BUSINESS")

    elif t.style == "hotel":
        c.setFont("Helvetica-Bold", 17)
        c.drawString(1.0 * inch, y, vendor)
        y -= 15
        c.setFont("Helvetica", 9)
        c.drawString(1.0 * inch, y, "GUEST FOLIO / STATEMENT OF ACCOUNT")
        y -= 28
        c.setFont("Helvetica", 9.5)
        c.drawString(1.0 * inch, y, f"Guest: {t.employee}")
        c.drawRightString(w - 1.0 * inch, y, f"Folio: {random.randint(200000, 299999)}")
        y -= 13
        c.drawString(
            1.0 * inch, y, f"Check-in: {(d - timedelta(days=2)).strftime('%m/%d/%Y')}"
        )
        c.drawRightString(w - 1.0 * inch, y, f"Check-out: {d.strftime('%m/%d/%Y')}")
        y -= 20
        c.setFont("Helvetica-Bold", 9)
        c.drawString(1.0 * inch, y, "DATE")
        c.drawString(2.0 * inch, y, "DESCRIPTION")
        c.drawRightString(w - 1.0 * inch, y, "AMOUNT")
        y -= 6
        c.line(1.0 * inch, y, w - 1.0 * inch, y)
        y -= 14
        for i, it in enumerate(t.items):
            c.setFont("Helvetica", 9.5)
            c.drawString(
                1.0 * inch, y, (d - timedelta(days=2 - min(i, 2))).strftime("%m/%d")
            )
            c.drawString(2.0 * inch, y, it.label)
            c.drawRightString(w - 1.0 * inch, y, _money(it.amount, cur))
            y -= 13
        y -= 4
        c.line(4.5 * inch, y, w - 1.0 * inch, y)
        y -= 16
        c.setFont("Helvetica-Bold", 11)
        c.drawRightString(w - 2.2 * inch, y, "BALANCE DUE:")
        c.drawRightString(w - 1.0 * inch, y, _money(total, cur))

    elif t.style == "airline":
        c.setFont("Helvetica-Bold", 16)
        c.drawString(1.0 * inch, y, vendor)
        y -= 14
        c.setFont("Helvetica", 10)
        c.drawString(1.0 * inch, y, "ELECTRONIC TICKET RECEIPT")
        y -= 26
        c.setFont("Helvetica", 9.5)
        c.drawString(1.0 * inch, y, f"Passenger: {t.employee.upper()}")
        y -= 13
        c.drawString(
            1.0 * inch,
            y,
            f"Confirmation: {''.join(random.choices('ABCDEFGHJKLMNPQRSTUVWXYZ23456789', k=6))}",
        )
        y -= 13
        c.drawString(1.0 * inch, y, f"Issue Date: {d.strftime('%d %b %Y').upper()}")
        y -= 22
        c.setFont("Helvetica-Bold", 9.5)
        c.drawString(1.0 * inch, y, "ITINERARY")
        y -= 15
        c.setFont("Helvetica", 9.5)
        c.drawString(1.0 * inch, y, "PDX -> DEN   Flight 1142   Depart 07:15")
        y -= 13
        c.drawString(1.0 * inch, y, "DEN -> PDX   Flight 1187   Depart 18:40")
        y -= 22
        for it in t.items:
            c.drawString(1.0 * inch, y, it.label)
            c.drawRightString(w - 1.0 * inch, y, _money(it.amount, cur))
            y -= 13
        y -= 4
        c.line(4.5 * inch, y, w - 1.0 * inch, y)
        y -= 16
        c.setFont("Helvetica-Bold", 11)
        c.drawRightString(w - 2.2 * inch, y, "TOTAL FARE:")
        c.drawRightString(w - 1.0 * inch, y, _money(total, cur))

    elif t.style == "rideshare":
        c.setFont("Helvetica-Bold", 18)
        c.drawString(1.0 * inch, y, vendor)
        y -= 30
        c.setFont("Helvetica", 12)
        c.drawString(1.0 * inch, y, f"Thanks for riding, {t.employee.split()[0]}")
        y -= 20
        c.setFont("Helvetica", 9.5)
        c.drawString(1.0 * inch, y, d.strftime("%B %d, %Y  •  7:42 PM"))
        y -= 28
        c.setFont("Helvetica-Bold", 24)
        c.drawString(1.0 * inch, y, _money(total, cur))
        y -= 30
        c.line(1.0 * inch, y, w - 1.0 * inch, y)
        y -= 18
        for it in t.items:
            c.setFont("Helvetica", 9.5)
            c.drawString(1.0 * inch, y, it.label)
            c.drawRightString(w - 1.0 * inch, y, _money(it.amount, cur))
            y -= 14

    else:  # invoice
        c.setFont("Helvetica-Bold", 16)
        c.drawString(1.0 * inch, y, vendor)
        c.setFont("Helvetica", 9)
        c.drawRightString(w - 1.0 * inch, y, "INVOICE")
        y -= 16
        c.setFont("Helvetica", 9)
        c.drawString(1.0 * inch, y, "Keizersgracht 241, 1016 EA Amsterdam, NL")
        y -= 11
        c.drawString(1.0 * inch, y, "VAT ID: NL8234.11.892.B01")
        y -= 26
        c.setFont("Helvetica", 9.5)
        c.drawString(1.0 * inch, y, f"Invoice No: INV-{random.randint(80000, 89999)}")
        c.drawRightString(w - 1.0 * inch, y, f"Invoice Date: {d.strftime('%Y-%m-%d')}")
        y -= 13
        c.drawString(1.0 * inch, y, "Bill To: Meridian Consulting Group")
        y -= 24
        c.line(1.0 * inch, y, w - 1.0 * inch, y)
        y -= 16
        for it in t.items:
            c.setFont("Helvetica", 9.5)
            c.drawString(1.0 * inch, y, it.label)
            c.drawRightString(w - 1.0 * inch, y, _money(it.amount, cur))
            y -= 14
        if t.tax:
            c.drawRightString(w - 2.2 * inch, y, "VAT 21%:")
            c.drawRightString(w - 1.0 * inch, y, _money(t.tax, cur))
            y -= 14
        y -= 4
        c.line(4.5 * inch, y, w - 1.0 * inch, y)
        y -= 16
        c.setFont("Helvetica-Bold", 11)
        c.drawRightString(w - 2.2 * inch, y, "TOTAL DUE:")
        c.drawRightString(w - 1.0 * inch, y, _money(total, cur))

    c.showPage()
    c.save()


def rasterize(path: Path) -> None:
    """Flatten a PDF to an image-only PDF: the text layer is destroyed, so any
    tool must fall back to OCR. Roughly a quarter of real receipts arrive this
    way, as a phone photo or a scan."""
    doc = pdfium.PdfDocument(str(path))
    page = doc[0]
    w_pt, h_pt = page.get_size()
    image = page.render(scale=150 / 72).to_pil()  # 150dpi, a typical scan
    page.close()
    doc.close()  # release the handle before overwriting the same path

    c = canvas.Canvas(str(path), pagesize=(w_pt, h_pt))
    c.drawImage(ImageReader(image), 0, 0, width=w_pt, height=h_pt)
    c.showPage()
    c.save()


# --------------------------------------------------------------------------
# Transaction construction
# --------------------------------------------------------------------------


def build_clean(i: int, d: date) -> Txn:
    cat = random.choice(list(VENDORS))
    vendor, style = random.choice(VENDORS[cat])
    emp, appr = random.choice(EMPLOYEES)

    if style in ("restaurant",):
        items = [
            LineItem(x, round(random.uniform(9, 26), 2))
            for x in random.sample(
                [
                    "Grilled salmon",
                    "House salad",
                    "Ribeye 12oz",
                    "Pasta primavera",
                    "Iced tea",
                    "Espresso",
                    "Soup du jour",
                    "Club sandwich",
                ],
                k=random.randint(2, 3),
            )
        ]
        sub = sum(i.amount for i in items)
        tax = round(sub * 0.0875, 2)
        tip = round(sub * random.choice([0.18, 0.20]), 2)
        total = round(sub + tax + tip, 2)
    elif style == "hotel":
        n = random.randint(2, 3)
        rate = round(random.uniform(139, 239), 2)
        items = [LineItem(f"Room charge - night {k + 1}", rate) for k in range(n)]
        items.append(LineItem("Occupancy tax", round(rate * n * 0.14, 2)))
        total = round(sum(i.amount for i in items), 2)
        tax = 0.0
        tip = 0.0
    elif style == "airline":
        base = round(random.uniform(210, 520), 2)
        items = [
            LineItem("Base fare", base),
            LineItem("Taxes and carrier fees", round(base * 0.21, 2)),
        ]
        total = round(sum(i.amount for i in items), 2)
        tax = tip = 0.0
    elif style == "rideshare":
        base = round(random.uniform(11, 44), 2)
        items = [
            LineItem("Trip fare", base),
            LineItem("Booking fee", 2.75),
            LineItem("Tip", round(base * 0.15, 2)),
        ]
        total = round(sum(i.amount for i in items), 2)
        tax = tip = 0.0
    elif style == "invoice":
        base = round(random.choice([49.0, 99.0, 149.0, 249.0]), 2)
        items = [LineItem("Team plan - monthly subscription", base)]
        tax = round(base * 0.21, 2)
        total = round(base + tax, 2)
        tip = 0.0
    else:  # retail / taxi
        items = [
            LineItem(x, round(random.uniform(4, 48), 2))
            for x in random.sample(
                [
                    "Copy paper 5-ream",
                    "Whiteboard markers",
                    "Notebooks (12pk)",
                    "Toner cartridge",
                    "Desk organizer",
                    "Sticky notes",
                ],
                k=random.randint(1, 3),
            )
        ]
        sub = sum(i.amount for i in items)
        tax = round(sub * 0.0875, 2)
        total = round(sub + tax, 2)
        tip = 0.0

    return Txn(
        txn_id=f"TX-{1000 + i}",
        employee=emp,
        sheet_date=d,
        sheet_vendor=vendor,
        category=cat,
        sheet_amount=total,
        sheet_currency="USD",
        purpose=random.choice(PURPOSES),
        approver=appr,
        receipt_file=f"TX-{1000 + i}.pdf",
        items=items,
        tax=tax,
        tip=tip,
        style=style,
        pdf_total=total,
    )


def plant_defects(txns: list[Txn]) -> None:
    """Mutate specific rows into the classic audit exceptions. Every mutation is
    recorded on the Txn so EXPECTED-FINDINGS.md stays in sync with reality."""

    def pick(idx: int) -> Txn:
        return txns[idx]

    # 1. Missing receipt entirely
    t = pick(3)
    t.receipt_file = None
    t.sheet_amount = 218.44  # above the 75.00 documentary-evidence floor
    # A vendor unique to this row, so this defect cannot get swept into the
    # split-transaction cluster below. Each planted defect should test one rule.
    t.sheet_vendor = "Northgate Conference Center"
    t.category = "Travel"
    t.defects.append(
        "MISSING_RECEIPT: 218.44 claimed with no support document, above the 75.00 floor"
    )

    # Deliberately NOT an exception. A small expense with no receipt is allowed
    # when the row carries amount, date, place and purpose. A tool that flags
    # this anyway floods the queue and gets ignored.
    t = pick(27)
    t.receipt_file = None
    t.sheet_amount = 14.80
    t.category = "Meals"

    # 2. Amount mismatch, digit transposition (127.40 typed as 172.40)
    t = pick(5)
    t.sheet_amount = round(t.pdf_total + 45.00, 2)
    t.defects.append(
        f"AMOUNT_MISMATCH: sheet claims {t.sheet_amount:.2f}, receipt totals {t.pdf_total:.2f} "
        f"(overclaim of {t.sheet_amount - t.pdf_total:.2f})"
    )

    # 3. Amount mismatch, small underclaim (honest error, still an exception)
    t = pick(12)
    t.sheet_amount = round(t.pdf_total - 3.10, 2)
    t.defects.append(
        f"AMOUNT_MISMATCH: sheet claims {t.sheet_amount:.2f}, receipt totals {t.pdf_total:.2f} "
        f"(underclaim of {t.pdf_total - t.sheet_amount:.2f})"
    )

    # 4. Duplicate: two rows point at the same receipt file
    src, dup = pick(7), pick(19)
    dup.receipt_file = src.receipt_file
    dup.sheet_amount = src.sheet_amount
    dup.sheet_vendor = src.sheet_vendor
    dup.sheet_date = src.sheet_date
    # The duplicate row's support document IS src's PDF, so every receipt-side
    # field must mirror src. Otherwise the answer key describes a file that does
    # not exist and the extractor gets blamed for reading the real one.
    dup.pdf_total = src.pdf_total
    dup.pdf_vendor = src.resolved("vendor")
    dup.pdf_date = src.resolved("date")
    dup.pdf_currency = src.resolved("currency")
    dup.style = src.style
    dup.items = list(src.items)
    dup.tax, dup.tip = src.tax, src.tip
    dup.defects.append(
        f"DUPLICATE_RECEIPT: row reuses the same support document as {src.txn_id}"
    )
    # The tool cannot know which row is the original, so both must be reviewed.
    src.defects.append(
        f"DUPLICATE_RECEIPT: this support document is also cited by {dup.txn_id}"
    )

    # 5. Date mismatch: receipt predates the claimed date by 9 days
    t = pick(9)
    t.pdf_date = t.sheet_date - timedelta(days=9)
    t.defects.append(
        f"DATE_MISMATCH: sheet says {t.sheet_date:%Y-%m-%d}, receipt says {t.pdf_date:%Y-%m-%d}"
    )

    # 6. Vendor mismatch: receipt is from a different merchant entirely
    t = pick(14)
    t.pdf_vendor = "Lakeshore Wine & Spirits"
    t.style = "retail"
    t.items = [
        LineItem("Cabernet Sauvignon 750ml", 62.00),
        LineItem("Single malt 750ml", 88.00),
    ]
    t.tax = round(150.00 * 0.0875, 2)
    t.tip = 0.0
    t.pdf_total = round(150.00 + t.tax, 2)
    t.sheet_amount = t.pdf_total
    t.defects.append(
        f"VENDOR_MISMATCH: sheet says '{t.sheet_vendor}', receipt says '{t.pdf_vendor}'"
    )
    t.defects.append(
        "POLICY_ALCOHOL: receipt line items are alcohol, disallowed under policy"
    )

    # 7. Foreign currency: EUR receipt booked as USD at face value
    t = pick(16)
    t.style = "invoice"
    t.sheet_vendor = "Nordvane Analytics BV"
    t.pdf_currency = "EUR"
    t.items = [LineItem("Team plan - monthly subscription", 149.00)]
    t.tax = round(149.00 * 0.21, 2)
    t.pdf_total = round(149.00 + t.tax, 2)
    t.sheet_amount = t.pdf_total  # same number, wrong currency
    t.sheet_currency = "USD"
    t.defects.append(
        "CURRENCY_MISMATCH: receipt is denominated in EUR but the report books the "
        "identical figure as USD, with no FX conversion applied"
    )

    # 8. Weekend charge on a non-travel category
    t = pick(21)
    while t.sheet_date.weekday() < 5:
        t.sheet_date += timedelta(days=1)
    t.pdf_date = t.sheet_date
    t.category = "Office Supplies"
    t.defects.append(
        f"WEEKEND_CHARGE: {t.category} expense dated {t.sheet_date:%A %Y-%m-%d}"
    )

    # 9. Split transaction to stay under a 500.00 approval threshold
    a, b = pick(23), pick(24)
    b.employee, b.approver = a.employee, a.approver
    for part, amt in ((a, 487.50), (b, 462.25)):
        part.sheet_vendor = "Ashcroft Chophouse"
        part.pdf_vendor = "Ashcroft Chophouse"
        part.style = "restaurant"
        part.category = "Client Entertainment"
        part.sheet_date = a.sheet_date
        part.pdf_date = a.sheet_date
        sub = round(amt / 1.2675, 2)
        part.items = [LineItem("Private dining - prix fixe", sub)]
        part.tax = round(sub * 0.0875, 2)
        part.tip = round(amt - sub - part.tax, 2)
        part.pdf_total = amt
        part.sheet_amount = amt
    a.defects.append(
        "SPLIT_TRANSACTION: paired with the next row, same vendor and same day, each "
        "just under the 500.00 approval threshold, 949.75 combined"
    )
    b.defects.append("SPLIT_TRANSACTION: second half of the split above")

    # 10. Tip added after the receipt was printed
    t = pick(30)
    t.style = "restaurant"
    t.category = "Client Entertainment"
    t.sheet_vendor = "The Regency Room"
    t.pdf_vendor = None
    t.items = [
        LineItem("Prix fixe dinner", 96.00),
        LineItem("Prix fixe dinner", 96.00),
        LineItem("Sparkling water", 9.00),
    ]
    sub = sum(i.amount for i in t.items)
    t.tax = round(sub * 0.0875, 2)
    t.tip = 0.0  # tip line left blank, filled in by hand later
    t.pdf_total = round(sub + t.tax, 2)
    t.sheet_amount = round(t.pdf_total + 40.00, 2)
    t.defects.append(
        f"UNSUPPORTED_TIP: sheet total {t.sheet_amount:.2f} exceeds printed receipt total "
        f"{t.pdf_total:.2f} by 40.00, consistent with a handwritten tip"
    )

    # 11. Personal expense misfiled as business
    t = pick(33)
    t.style = "invoice"
    t.sheet_vendor = "Streamflix Entertainment"
    t.pdf_vendor = "Streamflix Entertainment"
    t.category = "Software"
    t.items = [LineItem("Premium plan - personal account", 22.99)]
    t.tax = 0.0
    t.pdf_total = 22.99
    t.sheet_amount = 22.99
    t.purpose = "Monthly subscription"
    t.defects.append(
        "PERSONAL_EXPENSE: streaming entertainment subscription claimed as business software"
    )

    # 12. Meal over the itemization threshold with no itemized detail
    t = pick(36)
    t.style = "restaurant"
    t.category = "Client Entertainment"  # limit 500, so only itemization fires
    t.items = [LineItem("Dinner", 268.40)]
    t.tax = 0.0
    t.tip = 0.0
    t.pdf_total = 268.40
    t.sheet_amount = 268.40
    t.defects.append(
        "MISSING_ITEMIZATION: meal of 268.40 exceeds the 75.00 itemization threshold but "
        "the receipt shows a single undifferentiated line"
    )

    # 13. Scanned, image-only receipts, no text layer at all
    for idx in (2, 11, 25, 38):
        if idx < len(txns) and txns[idx].receipt_file:
            txns[idx].scanned = True
            txns[idx].defects.append(
                "SCAN_ONLY (not an exception): image-only PDF with no text layer, "
                "exercises the OCR fallback tier"
            )


# --------------------------------------------------------------------------
# Workbook
# --------------------------------------------------------------------------


def write_workbook(txns: list[Txn], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    headers = [
        "Txn ID",
        "Employee",
        "Date",
        "Vendor",
        "Category",
        "Amount",
        "Currency",
        "Receipt File",
        "Business Purpose",
        "Approver",
    ]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="1F3864")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")

    for t in txns:
        ws.append(
            [
                t.txn_id,
                t.employee,
                t.sheet_date.strftime("%Y-%m-%d"),
                t.sheet_vendor,
                t.category,
                t.sheet_amount,
                t.sheet_currency,
                t.receipt_file or "",
                t.purpose,
                t.approver,
            ]
        )

    for col, width in zip("ABCDEFGHIJ", (10, 22, 12, 26, 20, 11, 10, 16, 28, 14)):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        row[0].number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    wb.save(str(path))


def write_expectations(txns: list[Txn], path: Path) -> None:
    lines = [
        "# Expected findings in the sample data",
        "",
        "This file is generated by `tools/make_sample_data.py` alongside the sample set.",
        "It is the answer key. Every defect below was planted on purpose, so you can",
        "check whether an audit tool actually finds real problems.",
        "",
        "Rows not listed here are clean and should produce no exception.",
        "",
        "| Txn ID | Finding |",
        "| --- | --- |",
    ]
    n = 0
    for t in txns:
        for d in t.defects:
            lines.append(f"| {t.txn_id} | {d} |")
            n += 1
    real = sum(1 for t in txns for d in t.defects if not d.startswith("SCAN_ONLY"))
    lines[5:5] = [
        f"**{real} genuine exceptions** are planted across {len(txns)} transactions.",
        "",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# --------------------------------------------------------------------------


def write_truth(txns: list[Txn], path: Path) -> None:
    """Machine-readable answer key, so tests assert against the generator's
    intent rather than against a previous run's output. Prose alone rots."""
    import json

    rows = []
    for t in txns:
        rows.append(
            {
                "txn_id": t.txn_id,
                "employee": t.employee,
                "category": t.category,
                "purpose": t.purpose,
                "approver": t.approver,
                "receipt_file": t.receipt_file,
                "scan_only": t.scanned,
                # what the receipt PDF genuinely says
                "receipt_total": round(t.pdf_total, 2)
                if t.pdf_total is not None
                else None,
                "receipt_currency": t.resolved("currency"),
                "receipt_vendor": t.resolved("vendor"),
                "receipt_date": t.resolved("date").isoformat(),
                # what the expense report claims
                "sheet_amount": round(t.sheet_amount, 2),
                "sheet_currency": t.sheet_currency,
                "sheet_vendor": t.sheet_vendor,
                "sheet_date": t.sheet_date.isoformat(),
                # planted exceptions, by rule code
                "expected_codes": sorted(
                    {
                        d.split(":")[0]
                        for d in t.defects
                        if not d.startswith("SCAN_ONLY")
                    }
                ),
            }
        )
    path.write_text(json.dumps({"transactions": rows}, indent=2), encoding="utf-8")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--count",
        type=int,
        default=40,
        help="number of transactions (default 40; try 350 for a scale test)",
    )
    args = ap.parse_args()

    if args.count < 40:
        raise SystemExit(
            "--count must be at least 40; the planted defects need 40 rows"
        )

    random.seed(SEED)
    RECEIPTS.mkdir(parents=True, exist_ok=True)
    for old in RECEIPTS.glob("*.pdf"):
        old.unlink()

    start = date(2026, 7, 1)
    txns = [
        build_clean(i, start + timedelta(days=(i * 17) % 31)) for i in range(args.count)
    ]
    plant_defects(txns)

    written = 0
    for t in txns:
        if not t.receipt_file:
            continue
        target = RECEIPTS / t.receipt_file
        if target.exists():
            continue  # duplicate row deliberately points at an existing file
        render_receipt(t, target)
        if t.scanned:
            rasterize(target)
        written += 1

    # Manifest so the browser demo can load the sample set without a
    # directory listing, which static hosts do not provide.
    import json as _json

    names = sorted({t.receipt_file for t in txns if t.receipt_file})
    (OUT / "manifest.json").write_text(
        _json.dumps({"receipts": names, "sheet": "expense-report.xlsx"}, indent=2),
        encoding="utf-8",
    )

    write_workbook(txns, OUT / "expense-report.xlsx")
    write_expectations(txns, OUT / "EXPECTED-FINDINGS.md")
    write_truth(txns, OUT / "ground-truth.json")

    real = sum(1 for t in txns for d in t.defects if not d.startswith("SCAN_ONLY"))
    print(f"transactions : {len(txns)}")
    print(f"receipt PDFs : {written} ({sum(1 for t in txns if t.scanned)} image-only)")
    print(f"exceptions   : {real} planted")
    print(f"output       : {OUT}")


if __name__ == "__main__":
    main()
